# Update Date:  2001.03.24
# Author:       Grothendieck
# Version:      1.0
import os
import sys
import time
import re
import win32com.client

def main(path,config):
    SPfiles = ['cj0.sp','cj.sp','iv_forward.sp','iv_reverse.sp']
    lisfiles = ['cj0.lis','cj.lis','iv_forward.lis','iv_reverse.lis']

    def splitline(str):
        return str.rstrip('\n').split('|')[1].rstrip(' ').lstrip(' ')

    def ExtractCfg(cfg):
        dic= {}    
        lines = open(cfg,'r').readlines()
        for i,line in enumerate(lines):
            if 'model_lib_file_name' in line: 
                dic['lib_name'] = splitline(line)
            elif 'library_path' in line:
                dic['lib_path'] = splitline(line)
            elif 'before_merge or final_model' in line:
                dic['type'] = splitline(line)
            elif 'CJ0_voltage' in line:
                dic['cj0.sp'] = splitline(line)
            elif 'CJ_voltage' in line:
                dic['cj.sp'] = splitline(line)
            elif 'IV_forward' in line:
                dic['iv_forward.sp'] = splitline(line)
            elif 'IV_reverse' in line:
                dic['iv_reverse.sp'] = splitline(line)
            elif 'corner_definiton' in line:
                corners = splitline(line).split(' ')
                while '' in corners: corners.remove('')
                dic['corner'] = corners
            elif 'device_list' in line:
                devices = []
                for subline in lines[i+1:]:
                    if '-------' in subline: break
                    devices.append(subline.rstrip('\n'))
                while '' in devices: devices.remove('')
                dic['devices'] = devices    
        return dic

    def terminal(sp,cfg):
        # generate netlist
        lib = cfg['lib_path']+'/'+cfg['lib_name']
        corner1,corner2,corner3 = cfg['corner'][0],cfg['corner'][1],cfg['corner'][2]
        f = open(path + '/netlist/' + sp,'w')
        f.write("""%s Simulation
.OPTIONS beep=1 WL DCCAP=1 POST=2 INGOLD=2 measdgt=5 TNOM=25 POST_VERSION=9007 gmindc=1e-14 co=256 scale=1
.protect
.lib '%s' %s
.unprotect
.TEMP 25 -40 125\n
"""%(sp,lib,corner1))
        if sp in ['cj.sp','cj0.sp']:
            f.write(".param Vdio= %s"%cfg[sp]+'\n')
            f.write(".AC POI 1 100000"+'\n\n')
            for i,device in enumerate(cfg['devices']):
                index = str(i+1).rjust(2,'0')
                f.write("d%s D%s 0 %s"%(index,index,device)+'\n')
                f.write("Vac%s D%s DT%s dc=0 ac=0.1"%(index,index,index)+'\n')
                f.write("Vdc%s DT%s 0 dc='Vdio'"%(index,index)+'\n')
                f.write(".meas ac i_in%s find II(Vac%s) AT=100000"%(index,index)+'\n')
                f.write(".meas ac dio_cv%s param='-i_in%s/0.1/(2*3.14159265358*100000)*1e12'"%(index,index)+'\n\n')
        elif sp in ['iv_forward.sp','iv_reverse.sp']:
            f.write('.param Vdio= %s'%cfg[sp]+'\n')
            f.write('.param vgsweep= 0'+'\n')
            f.write('.dc vgsweep 0 Vdio Vdio'+'\n\n')
            for i,device in enumerate(cfg['devices']):
                index = str(i+1).rjust(2,'0')
                f.write("d%s in%s Db%s %s"%(index,index,index,device)+'\n')
                f.write("Vin%s in%s 0 vgsweep"%(index,index)+'\n')
                f.write("VDb%s Db%s 0 0"%(index,index)+'\n')
                f.write(".meas dc i_in%s find i(Vin%s) when v(in%s,0)=Vdio"%(index,index,index)+'\n')
                f.write(".meas dc dio_iv%s param='-i_in%s'"%(index,index)+'\n\n')
        f.write("""
.alter
.protect
.lib '%s' %s
.unprotect

.alter
.protect
.lib '%s' %s
.unprotect"""%(lib,corner2,lib,corner3))
        f.write('\n\n.end')
        f.close()

    cfg = ExtractCfg(path + '/' + config)   

    if os.path.exists(path+'/netlist'): pass
    else: os.mkdir(path+'/netlist') 

    for sp in SPfiles:
        terminal(sp,cfg)        

    def check_hspice():
        # monitor the hspice.exe
        WMI = win32com.client.GetObject('winmgmts:')
        processCodeCov = WMI.ExecQuery('select * from Win32_process where Name="hspice.exe"')
        if len(processCodeCov) > 0:
            return True
        else:
            return False
    
    print('\n Hspice 仿真中，请等待...')
    
    for sp in SPfiles:
        input = r'netlist\%s'%sp
        output = r'netlist\%s'%sp.split('.')[0]
        # command to excute the hspice.exe
        os.system( r'hspice -i %s -o %s'%(input,output) )
        while(check_hspice()):
            time.sleep(1)
        print('\n %s simulation done.'%sp)

    print('\n 仿真完毕，开始抽取数据...')

    def SimulationData(lis):
        if lis in ['cj.lis','cj0.lis']:
            seed = 'dio_cv'
        elif lis in ['iv_forward.lis','iv_reverse.lis']:
            seed = 'dio_iv'
        content = open(path+ '/netlist/' + lis,'r').read()
        ans = re.findall(seed + '\d+\s*=\s*(-?\d\.\d+(e|E)(-|\+)\d+)',content)    
        return [eval(x[0]) for x in ans]
    
    if os.path.exists(path+'/'+table): pass
    else: download(table)

    import openpyxl
    wb = openpyxl.load_workbook(path + '/Diode Crosscheck.xlsx')
    for lis in lisfiles:
        ws = wb[lis.split('.')[0]]
        if cfg['type'] == 'before_merge': colstart = 2
        elif cfg['type'] == 'final_model': colstart = 11
        for i,device in enumerate(cfg['devices']):
            ws.cell(row = 4+i, column = 1).value = device.split(' ')[0]
        for i,data in enumerate(SimulationData(lis)):
            row = 4 + i%len(cfg['devices'])
            col = i//len(cfg['devices']) + colstart
            ws.cell(row = row, column = col).value = data
    wb.save(path + '/Diode Crosscheck.xlsx')

    print('\n 结果已保存至 .xlsx 中，请查看')

if 1:	
    print('''
        *********************************************
                  Diode Model Cross Check
        *********************************************
        ''')
    
    path = os.getcwd()
    #path = sys.path[0]

    #-------------------------------------------- Config File -----------------------------------------------------
    remoteDisk = r'\\tdtbde\TBDE\CompactModel2\NHD_Model2\Tools\Smic.go Repository'	
    config = 'Diode_Crosscheck_Config.ini'
    table = 'Diode Crosscheck.xlsx'

    def download(f):
        import shutil
        shutil.copyfile(remoteDisk + '/' + f, path + '/' + f)
    
    AlreadyExist = os.path.exists(path + '/' + config)
    NoneExist = not AlreadyExist
    
    if NoneExist:
        ans = input(' 是否下载配置文件 ? \n\n y or n ? : ')
        download(config) if ans == 'y' else exit()
    elif AlreadyExist:
        ans = input('\n 检测到你已经有了配置文件, 是否需要重新下载一份？\n\n y or n ? : ')
        download(config) if ans == 'y' else print('\n')
    
    ans = input(' 请对配置文件进行修改并保存，若无需修改，或已修改完毕，则按 y 继续。 \n\n 是否继续? : ')
    main(path,config) if ans == 'y' else exit()
    #--------------------------------------------------------------------------------------------------------------------

if 0:
    if __name__ == '__main__':
        pass
    
