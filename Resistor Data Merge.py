# Update Date:  2016.11.23
# Author:       Grothendieck
# Version:      1.0.0   :   Basic function completed.
import os
import sys
import xlrd
import openpyxl

def main(folder,manual):

    data = xlrd.open_workbook(manual)
    sheet_list = [line.name for line in data.sheets()]
    if 'Resistor' in sheet_list: t = data.sheet_by_name('Resistor');res='Resistor'
    elif 'resistor' in sheet_list: t = data.sheet_by_name('resistor');res='resistor'
    else: print('\n 请将Manual中Resistor部分的sheet改名。')

    nums = t.nrows

    start = {'START:','START','START：','start:','Start:'}
    end   = {'END:','END','END：','end:','End:'}
    W     = {'W','w'}
    L     = {'L','l'}
    No    = {'N','n','No','NO','no','N0','n0'}

    block = []
    for i in range(nums):
        line = t.row_values(i)
        
        if len(start & set(line)) != 0:
            istart = i
            # the column where the 'START' is:
            colS = line.index(list(start & set(line))[0])
            
        elif len(end & set(line)) != 0:
            iend = i
            # count sub testkey numbers:
            tn = iend - istart - 4
            # testkey numbers:
            testkey = t.row_values(istart+1)[colS]
            # type description:
            type = t.row_values(istart+2)[colS]
            # record testkey block's location: [Testkey,rowSTART,rowEND,colSTART,type]:
            block.append([testkey,istart,iend,colS,type])

            print(' ',testkey,' : ',type,'\t 计数 : ',tn)

    ans = input('''
 请检查各型电阻的Testkey数量是否正确，无问题则按'y'继续。
 如有问题，按'n'退出，并请检查manual中的'START'或'END'有无缺少。
 
 是否继续？y or n ? : ''')
    
    if ans == 'y':pass;print('\n')
    else: exit()
    #--------------------------------------------extract all testkeys' information------------------------------------------------
    listdir = os.listdir(folder)
    # include testkey/W/L/No/dataFiles:
    all = []
    for b in block:
        #------------------------------------------get W/L/No column-----------------------------------------
        # 3 rows under 'START'
        line = t.row_values(b[1]+3)                                             
        colS = b[3]
        # the column where the 'W' is
        # .................... 'L' ..
        # .................... 'No' .
        colW = line.index(list(W & set(line))[0])                               
        colL = line.index(list(L & set(line))[0])                               
        colN = line.index(list(No & set(line))[0])                              
        #----------------------------------------------------------------------------------------------------
        
        # sub testkey: a,b,c,d,e.....        
        for subtk in range(b[1]+4,b[2]):
            line = t.row_values(subtk)

            dic = {}
            # like: MA031a = MA031 + a :
            TK = b[0] + line[colS]
            dic['Testkey'] = TK
            # del '#' if it has:
            dic['type'] = b[4].lstrip('#')
            dic['W']    = line[colW]
            dic['L']    = line[colL]
            dic['No']   = line[colN]

            
            for file in listdir:
                # set them to upper style to compare:
                if TK.upper() in file.upper():
                    # the 11th row in data file is 'Instance':
                    instance = open(folder + '/' + file,'r').readlines()[10]
                    if 'T=25' in instance: dic['T25'] = file
                    elif 'T=-40' in instance: dic['T-40'] = file
                    elif 'T=125' in instance: dic['T125'] = file
            # dic = {'Testkey':xx,'type':xx,'W':xx,'L':xx,'No':xx,'T25':xx,'T-40':xx,'T125':xx}
            all.append(dic)
    #-----------------------------------------------------------------------------------------------------------------------------



    #--------------------------------------------merge every testkeys' data to table----------------------------------------------

    excel = openpyxl.Workbook()
    ws = excel.active
    ws.title = 'Manual'

    Manual = openpyxl.load_workbook(manual)[res]
    for row in Manual.rows:
        ws.append([x.value for x in row])
    del Manual

    type = set([x['type'] for x in all])
    title = ['Testkey','Type','W','L','Vin','V(25)','R(25)','V(40)','R(-40)','V(125)','R(125)']

    for i in type:
        # built sheets by all of type names
        excel.create_sheet(i)
        t2 = excel.get_sheet_by_name(i)
        for i,x in enumerate(title):
            # write in title
            t2.cell(row = 1,column = i+1).value = x

    def extract(word,content):
        lineStr = list(filter(lambda x: word in x, content)).pop()
        lineNum = content.index(lineStr)
        return lineNum

    #-------------------------------------------extract function for Vin/V/R--------------------------------------------------
    def VR(file):
        data = open(folder + '/' + file,'r').readlines()
        lens = len(data)
        # where the Rs line is:
        rs = extract('Vr,Rs',data)
        Vin,R = [],[]
        for line in data[rs+1:lens]:
            # split data of ',', transform to float format
            obj = list(map(float,line.split(',')))
            Vin.append(obj[0])
            R.append(obj[1])
        try:
            # where the Vv2,Vv1,Ir line is:
            v2 = extract('Vr,Vv2',data)
            v1 = extract('Vr,Vv1',data)
            Ir = extract('Vr,Ir',data)
            # select one pair data to test:
            obj2 = list(map(float,data[v2+1].split(',')))
            obj1 = list(map(float,data[v1+1].split(',')))
            # check which is V not I:
            if abs(obj2[0] - obj2[1]) < abs(obj1[0] - obj1[1]):
                begin = v2 + 1; end = v1
            else: begin = v1 +1; end = Ir

            V =[]
            for line in data[begin:end]:
                obj = list(map(float,line.split(',')))
                V.append(obj[1])
        except:
            # for 2T measurement,there are no Vv2 & Vv1:
            V = Vin

        return {'Vin':Vin,'V':V,'R':R}
    #---------------------------------------------------------------------------------------------------------------------------
    
    for tk in all:

        row = 1
        # find the correspond sheet by their type name:
        t2 = excel[tk['type']]
        # check how many rows does the sheet has yet:
        bottom = len(t2.rows)
        if 'T25' not in tk:
            print(' %s    :\t 无25度数据，跳过'%tk['Testkey'])
            pass
        else:
            Vin = VR(tk['T25'])['Vin']
            V25 = VR(tk['T25'])['V']
            R25 = VR(tk['T25'])['R']
            
            if 'T-40' not in tk:
                print(' %s    :\t 无-40度数据，此温度跳过'%tk['Testkey'])
                V40 = False 
            else:
                V40 = VR(tk['T-40'])['V']
                R40 = VR(tk['T-40'])['R']
                if len(V40)!=len(V25):
                    V40 = False
                    print(' %s    :\t -40度采样数与25度不相等，此温度跳过'%tk['Testkey'])
                    
            if 'T125' not in tk:
                print(' %s    :\t 无125度数据，此温度跳过'%tk['Testkey'])
                V125 = False 
            else:
                V125 = VR(tk['T125'])['V']
                R125 = VR(tk['T125'])['R']
                if len(V125)!=len(V25):
                    V125 = False
                    print(' %s    :\t 125度采样数与25度不相等，此温度跳过'%tk['Testkey'])


            t2.cell(row = bottom + row , column = 1).value = tk['Testkey']
            for i,v in enumerate(Vin):
                for j,x in enumerate(['type','W','L']):
                    # write in some information
                    t2.cell(row = bottom + row , column = j + 2).value = tk[x]

                t2.cell(row = bottom + row , column = 5).value = Vin[i]
                t2.cell(row = bottom + row , column = 6).value = V25[i]
                t2.cell(row = bottom + row , column = 7).value = R25[i]

                if V40 is False:
                    pass
                else:
                    t2.cell(row = bottom + row , column = 8).value = V40[i]
                    t2.cell(row = bottom + row , column = 9).value = R40[i]

                if V125 is False:
                    pass
                else:
                    t2.cell(row = bottom + row , column = 10).value = V125[i]
                    t2.cell(row = bottom + row , column = 11).value = R125[i]

                row += 1
            print(' %s    :\t 已完成'%tk['Testkey'])
    file = os.path.dirname(folder) + '/' + 'Resistor_Data_From_Raw_Data.xlsx'
    excel.save(file)
    #-----------------------------------------------------------------------------------------------------------------------------
    print('\n 所有数据已合并完成至 Resistor_Data_From_Raw_Data.xlsx 中，请查看。')

#---------------------------------------------------------main function end-------------------------------------------------------
if 1:
    path = os.getcwd()
    #path = sys.path[0]

    print('''
        **************************************
          Resistor Data Merge From Raw Data
        **************************************
        ''')

    flist = os.listdir(path)
    dic = {str(i+1):f for i,f in enumerate(flist)}

    for i,f in enumerate(os.listdir(path)):
        print(' ',i+1,'> ',f)

    print('\n 选择包含所有 Resistor Data 的文件夹:')
    ans1 = input('\n select : ')
    while ans1 not in dic:
        ans1 = input(' Please re-select : ')
    print('\n Your choice :\t\t%s'%dic[ans1])

    print('\n 选择包含 Manual 的表格文件:')
    ans2 = input('\n select : ')
    while ans2 not in dic:
        ans2 = input(' Please re-select : ')
    print('\n Your choice :\t\t%s'%dic[ans2])
    print('\n')

    folder,manual = path+'/'+dic[ans1], path+'/'+dic[ans2]
    main(folder,manual)
    exit()

if 0:
    if __name__=='__main__':
        path = sys.path[0]
        folder = path + '\All2'
        manual = path + '\Resistor Manual.xlsx'
        main(folder,manual)