# Update Date:	2016.11.13
# Author:		Grothendieck
# Version:		1.0.1	:	
#				1.0.0	:	
import os,sys
import shutil
import openpyxl
import xlrd,xlwt

def main(folder):

	# write in all data files' name to the excel
	def flistwrite(excel):
		filelist = os.listdir(folder)

		wb = openpyxl.load_workbook(excel) 
		sheet = wb.get_sheet_by_name('Varactor')
		
		row = 0
		for i,fname in enumerate(filelist):
			col = i % 6
			if col == 0:
				row += 1 
			sheet.cell(row = 19 + row, column = 11 + col).value = fname  # start from 20th row, 11th column 
		wb.save(excel)	

	remoteDisk = r'\\tdtbde\TBDE\CompactModel2\NHD_Model2\Tools\Smic.go Repository'
	Table = 'Varactor Data Merge.xlsx'	
	FileOnServer = remoteDisk + '/' + Table
	FileOnLocal  = os.path.dirname(folder) + '/' + Table
	AlreadyExist = os.path.exists(FileOnLocal)
	NoneExist 	 = not AlreadyExist
	
	if NoneExist:
		shutil.copyfile(FileOnServer, FileOnLocal)
		flistwrite(FileOnLocal)
		print('''
 'Varactor Data Merge.xlsx'已下载到当前文件夹，请打开模板，填入尺寸参数。
 并将所有文件名移动到对应位置，完成后请保存。''')
		
	if AlreadyExist:
		print("\n 检测到当前文件夹已有'Varactor Data Merge.xlsx'表格，是否需要重新下载？")
		ans = input('\n y or n ? : ')
		if ans == 'y':
			shutil.copyfile(FileOnServer,FileOnLocal)
			flistwrite(FileOnLocal)
			print('''
 'Varactor Data Merge.xlsx'已下载到当前文件夹，请打开模板，填入尺寸参数。
 并将所有文件名移动到对应位置，完成后请保存。''')
		else:
			pass

	ans = input('\n 修改好表格后，确认是否继续:\n\n y or n ? : ');print('\n')
	if ans == 'y': pass	
	else: input('\n press Enter to exit.'); exit()					



	#---------------------------------------------------generate a new table to display all data---------------------------------

	'''
	excel1 : Varactor Data Merge.xlsx
	excel2 : final excel for model exctration
	'''

	alignment	=	xlwt.Alignment()
	alignment.horz	=	xlwt.Alignment.HORZ_CENTER
	alignment.vert	=	xlwt.Alignment.VERT_CENTER
	#边框
	borders	=	xlwt.Borders()
	borders.left	=	1
	borders.right	=	1
	borders.top		=	1
	borders.bottom	=	1

	# 2种字体颜色
	font1	=	xlwt.Font()
	font1.name	=	'Arial' 
	font1.colour_index	=	0 # black

	font2	=	xlwt.Font()
	font2.name	=	'Arial' 
	font2.colour_index	=	1 # white

	font3	=	xlwt.Font()
	font3.name	=	'Arial' 
	font3.colour_index	=	4 # blue	
	
	# 6种填充颜色
	pattern1	=	xlwt.Pattern()
	pattern1.pattern	=	1
	xlwt.Pattern.SOLID_PATTERN #full fill
	pattern1.pattern_fore_colour	=	55 # grew

	pattern2	=	xlwt.Pattern()
	pattern2.pattern	=	1
	xlwt.Pattern.SOLID_PATTERN #full fill
	pattern2.pattern_fore_colour	=	57 # green

	pattern3	=	xlwt.Pattern()
	pattern3.pattern	=	1
	xlwt.Pattern.SOLID_PATTERN #full fill
	pattern3.pattern_fore_colour	=	26 # shallow purple

	pattern4	=	xlwt.Pattern()
	pattern4.pattern	=	1
	xlwt.Pattern.SOLID_PATTERN #full fill
	pattern4.pattern_fore_colour	=	43 # dark purple

	pattern5	=	xlwt.Pattern()
	pattern5.pattern	=	1
	xlwt.Pattern.SOLID_PATTERN #full fill
	pattern5.pattern_fore_colour	=	42 # shallow cyan

	pattern6	=	xlwt.Pattern()
	pattern6.pattern	=	1
	xlwt.Pattern.SOLID_PATTERN #full fill
	pattern6.pattern_fore_colour	=	27 # dark cyan


	# title1： white font, grey pattern 
	style1	=	xlwt.XFStyle() 
	style1.alignment	=	alignment
	style1.borders	=	borders
	style1.font		=	font2
	style1.pattern	=	pattern1

	# title2： white font, green pattern
	style2	=	xlwt.XFStyle() 
	style2.alignment	=	alignment
	style2.borders	=	borders
	style2.font		=	font2
	style2.pattern	=	pattern2

	# black font, shallow purple pattern
	style3	=	xlwt.XFStyle() 
	style3.alignment	=	alignment
	style3.borders	=	borders
	style3.font		=	font1
	style3.pattern	=	pattern3

	# black font, dark purple pattern
	style4	=	xlwt.XFStyle() 
	style4.alignment	=	alignment
	style4.borders	=	borders
	style4.font		=	font1
	style4.pattern	=	pattern4

	# black font, shallow cyan pattern
	style5	=	xlwt.XFStyle() 
	style5.alignment	=	alignment
	style5.borders	=	borders
	style5.font		=	font1
	style5.pattern	=	pattern5

	# black font, dark cyan pattern
	style6	=	xlwt.XFStyle() 
	style6.alignment	=	alignment
	style6.borders	=	borders
	style6.font		=	font1
	style6.pattern	=	pattern6

	# blue font, no pattern
	style7	=	xlwt.XFStyle() 
	style7.alignment	=	alignment
	style7.borders	=	borders
	style7.font		=	font3


	#------------------------------------read infomation form excel1--------------------------------

	excel1 = os.path.dirname(folder) + '/' + Table
	t1 = xlrd.open_workbook(excel1).sheet_by_name('Varactor') 

	testkey = t1.col_values(0)[:]
	testkey = testkey[2:]  													# start from the 3rd row
	while '' in testkey:
		testkey.remove('')

	total = []
	for i in range(2,2+len(testkey)):
		total.append(t1.row_values(i)[:])
	#----------------------------------------------end----------------------------------------------
		
		
		
	#-----------------------------------------write in excel2---------------------------------------

	excel2 = xlwt.Workbook(encoding='utf-8') 
	t2 = excel2.add_sheet('Sheet',cell_overwrite_ok=True) 
	title = ['Testkey','W','L','wr','lr','nf','Nx','Ny','Mr','Area','T','Voltage(V)','Si(pF)','cal(pF)','Si - Cal(fF)','Silicon_Normalized']	
	for i,x in enumerate(title):
		if i < 11: style = style1
		else: style = style2
		t2.write(0,i,x,style)

	def col(x):
		if len(x) == 1:
			col = ord(x)-ord('A')
		return col

	def Median(file):
		import numpy
		data = open(folder + '/' + file).readlines()
		lens = len(data) 
		Cgc = []
		for line in data[17:lens]:
			obj=list(map(float,line.split(',')))			# split data of ',', transform to float format
			Cgc.append(obj[1])
		result = numpy.median(Cgc)
		return result										# get the median value of calibate structure capacitance

		
	row = 1
	flag = 0               									# block color flag 
	for tk,x in enumerate(testkey): 						# testkey
		
		size = total[tk][0:10]

		for T in [25,-40,125]:
			if T == 25: device,calibate= col('K'),col('N')
			elif T == -40: device,calibate = col('L'),col('O')
			elif T == 125: device,calibate = col('M'),col('P')
			
			if total[tk][device] == '':								# if no device data file, pass
				print(' ',total[tk][0],'has no data in temperature',T,'\tpass')
				continue											# jump out the cicle
			else:								
				data = open(folder + '/' + total[tk][device]).readlines()
				lens = len(data) 
				Vgs,Cgc = [],[]
				for line in data[17:lens]:
					obj = list(map(float,line.split(',')))			# split data of ',', transform to float format
					Vgs.append(obj[0])
					Cgc.append(obj[1])
				if Vgs[0] < 0 and Vgs[-1] > 0:						# always from positive Vgs to negative Vgs
					Vgs.reverse()
					Cgc.reverse()
				
				if total[tk][calibate] == '':						# if no calibate data file, median value set to be 0
					median = 0
					print('\n ',total[tk][0],'in temperature',T,'has no calibate data, set to be 0')
				else:
					median = Median(total[tk][calibate])
				
				flag += 1
				if flag % 2 == 1: stylex = style3; styley = style4   
				elif flag % 2 == 0: stylex = style5; styley = style6 
				
				Vmax_row = row 										# this row is marked for normalizing			
				for i,x in enumerate(Vgs):
					for j,x in enumerate(size):
						t2.write(row,j,x,stylex)								# write in size infomation
						t2.write(row,col('K'),T,stylex)							# write in temperature
					t2.write(row,col('L'),Vgs[i],styley)						# write in Vgs
					t2.write(row,col('M'),Cgc[i],styley)						# write in Cgc
					t2.write(row,col('N'),median,styley)						# write in median
					t2.write(row,col('O'),xlwt.Formula('(M%d-N%d)*1000'%(row+1,row+1)),style7) 		# write in Si - Cal(fF)
					t2.write(row,col('P'),xlwt.Formula('O%d/$O$%d'%(row+1,Vmax_row+1)),style7)		# write in normalizing
					row += 1
	excel2.save(os.path.dirname(folder) + '/' + 'Varactor_Data_From_Raw_Data.xls')
	#---------------------------------------------end--------------------------------------------	
	print('\n 所有数据已合并完成至 Varactor_Data_From_Raw_Data.xls 中，请查看。')

#--------------------------------------------main function end---------------------------------------

if 1:
	path = os.getcwd()
	#path = sys.path[0]
		
	print('''
		**************************************
		  Varactor Data Merge From Raw Data
		**************************************
		''')	
	
		
	flist = os.listdir(path)	
	dic = {str(i+1):f for i,f in enumerate(flist)}	
	
	for i,f in enumerate(os.listdir(path)):
		print(' ',i+1,'> ',f)

	print('\n 选择包含所有 Varactor Data 的文件夹:')	
		
	ans = input('\n select : ')	
	while ans not in dic:
		ans = input(' Please re-select : ')	
	print('\n Your choice :\t\t%s'%dic[ans])			

	folder = path + '/' + dic[ans]
	
	main(folder)
	exit()
	
if 0:
	if __name__=='__main__':
		path = sys.path[0]
		folder = path+'\All'
		main(folder)


